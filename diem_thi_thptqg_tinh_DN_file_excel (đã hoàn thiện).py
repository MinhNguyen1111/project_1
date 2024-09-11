import os
import requests
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

def get_next_empty_row(ws):
    """Tìm dòng trống tiếp theo trong sheet."""
    for row in range(1, ws.max_row + 2):
        if all(cell.value is None for cell in ws[row]):
            return row
    return ws.max_row + 1

def auto_fit_columns(ws):
    """Tự động điều chỉnh kích thước cột để phù hợp với nội dung."""
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Thêm một chút không gian
        dim_holder[column] = ColumnDimension(ws, min=col, max=col, width=adjusted_width)
    ws.column_dimensions = dim_holder

# Đặt tên file Excel
excel_file = "diemthi2024.xlsx"

# Kiểm tra file Excel đã tồn tại chưa
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    start_row = get_next_empty_row(ws)
else:
    wb = Workbook()
    ws = wb.active
    start_row = 1

error_count = 0  # Biến đếm số lần bị lỗi liên tiếp
consecutive_error_count = 0  # Biến đếm số lần lỗi liên tiếp
max_consecutive_errors = 2  # Số lần lỗi liên tiếp tối đa trước khi dừng chương trình
batch_size = 100  # Kích thước của từng nhóm dữ liệu trước khi ghi vào file Excel

data_list = []  # Danh sách để lưu trữ dữ liệu trước khi ghi vào file Excel

for x in range(48000001, 49000000):
    try:
        data_url = "https://sgddt.dongnai.gov.vn/tracuu/export/" + str(x) + ".json"
        response = requests.get(data_url, allow_redirects=True, timeout=5)

        # Kiểm tra nếu yêu cầu thành công
        if response.status_code == 200:
            info = response.json()

            # Kiểm tra nếu dữ liệu hợp lệ
            if info and all(k in info for k in ('HO_TEN', 'SOBAODANH', 'DIEM_THI', 'NGAY_SINH', 'GIOI_TINH', 'CMND')):
                data_list.append({
                    "Họ tên": info['HO_TEN'],
                    "SBD": info['SOBAODANH'],
                    "Điểm": info['DIEM_THI'],
                    "Ngày sinh": info['NGAY_SINH'],
                    "Giới tính": info['GIOI_TINH'],
                    "Số cmnd": info['CMND']
                })
                print("Đang dò điểm thi của số báo danh: " + str(x))

                # Ghi dữ liệu thành từng đợt vào file Excel để tránh mất mát dữ liệu
                if len(data_list) >= batch_size:
                    df = pd.DataFrame(data_list)
                    for row in dataframe_to_rows(df, index=False, header=False):
                        ws.append(row)
                    data_list.clear()  # Xóa danh sách sau khi ghi

                consecutive_error_count = 0  # Reset lỗi liên tiếp

            else:
                print(f"Số báo danh {x} không có dữ liệu.")
                error_count += 1
                consecutive_error_count += 1

        else:
            print(f"Số báo danh {x} không có dữ liệu hoặc gặp lỗi khác (status_code: {response.status_code}).")
            error_count += 1
            consecutive_error_count += 1

    except requests.exceptions.RequestException as e:
        print(f"Số báo danh {x} gặp lỗi: {e}")
        error_count += 1
        consecutive_error_count += 1

    # Kiểm tra nếu số lần lỗi liên tiếp vượt quá giới hạn
    if consecutive_error_count > max_consecutive_errors:
        print("Quá nhiều lỗi liên tiếp, dừng chương trình khẩn cấp.")
        break

# Lưu dữ liệu còn lại nếu có
if data_list:
    df = pd.DataFrame(data_list)
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

# Tự động điều chỉnh kích thước cột
auto_fit_columns(ws)

# Lưu file Excel
wb.save(excel_file)
